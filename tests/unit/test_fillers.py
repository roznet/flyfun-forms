"""Tests for document fillers — PDF, French customs PDF, XLSX.

Each test generates a real document from the production template and verifies
the output is valid and contains expected data.
"""

from io import BytesIO
from pathlib import Path

import pytest
from pypdf import PdfReader
from openpyxl import load_workbook

from flightforms.api.models import GenerateRequest
from flightforms.fillers.pdf_filler import fill_pdf, _parse_date, _resolve_field_pattern
from flightforms.fillers.french_customs_filler import fill_french_customs, _suffix
from flightforms.fillers.xlsx_filler import fill_xlsx
from flightforms.registry import MappingRegistry
from tests.conftest import (
    MAPPINGS_DIR,
    TEMPLATES_DIR,
    StubAirportResolver,
    make_aircraft,
    make_crew_member,
    make_flight,
    make_passenger,
    make_pilot,
    make_second_passenger,
)


# ── Helper functions ──────────────────────────────────────────────────────────

class TestParseDatePdf:
    def test_simple(self):
        assert _parse_date("2099-06-15", "%d/%m/%Y") == "15/06/2099"

    def test_day_month_year(self):
        result = _parse_date("2099-01-03", "%d %B %Y")
        assert "3" in result and "January" in result and "2099" in result


class TestResolveFieldPattern:
    def test_zero_based(self):
        assert _resolve_field_pattern("FieldRow{i}", 0) == "FieldRow0"
        assert _resolve_field_pattern("FieldRow{i}", 2) == "FieldRow2"

    def test_one_based(self):
        assert _resolve_field_pattern("FieldRow{n}", 0) == "FieldRow1"
        assert _resolve_field_pattern("FieldRow{n}", 2) == "FieldRow3"

    def test_both(self):
        assert _resolve_field_pattern("F{i}_R{n}", 1) == "F1_R2"


class TestFrenchCustomsSuffix:
    def test_first_row(self):
        assert _suffix(0) == ""

    def test_second_row(self):
        assert _suffix(1) == "_2"

    def test_third_row(self):
        assert _suffix(2) == "_3"


# ── PDF filler (LSGS) ────────────────────────────────────────────────────────

class TestPdfFillerLSGS:
    @pytest.fixture
    def registry(self):
        return MappingRegistry(str(MAPPINGS_DIR), str(TEMPLATES_DIR))

    @pytest.fixture
    def resolver(self):
        return StubAirportResolver()

    def _generate(self, registry, resolver, flatten=False) -> bytes:
        mapping = registry.get_form("LSGS", "lsgs")
        request = GenerateRequest(
            airport="LSGS",
            form="lsgs",
            flight=make_flight(origin="ZZZZ", destination="LSGS"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
            passengers=[make_passenger()],
        )
        return fill_pdf(
            registry.get_template_path(mapping),
            mapping,
            request,
            resolver,
            flatten=flatten,
        )

    def test_output_is_valid_pdf(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        assert len(pdf_bytes) > 0
        reader = PdfReader(BytesIO(pdf_bytes))
        assert len(reader.pages) >= 1

    def test_output_flattened(self, registry, resolver):
        normal_bytes = self._generate(registry, resolver, flatten=False)
        flat_bytes = self._generate(registry, resolver, flatten=True)
        # Count annotations: flattened should have far fewer (widget annots removed)
        normal_reader = PdfReader(BytesIO(normal_bytes))
        flat_reader = PdfReader(BytesIO(flat_bytes))
        normal_count = sum(len(p.get("/Annots", [])) for p in normal_reader.pages)
        flat_count = sum(len(p.get("/Annots", [])) for p in flat_reader.pages)
        assert flat_count < normal_count

    def test_fields_filled(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, flatten=False)
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        # Registration should appear in the field values
        reg_field = fields.get("Text7")
        if reg_field:
            val = reg_field.get("/V", "")
            assert "ZZ-TST" in str(val)

    def test_direction_arrival(self, registry, resolver):
        """When destination == airport, direction should be inbound."""
        mapping = registry.get_form("LSGS", "lsgs")
        request = GenerateRequest(
            airport="LSGS",
            form="lsgs",
            flight=make_flight(origin="ZZZZ", destination="LSGS"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
        )
        # Just verify it doesn't crash; direction logic is internal
        result = fill_pdf(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0

    def test_direction_departure(self, registry, resolver):
        mapping = registry.get_form("LSGS", "lsgs")
        request = GenerateRequest(
            airport="LSGS",
            form="lsgs",
            flight=make_flight(origin="LSGS", destination="ZZZZ"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
        )
        result = fill_pdf(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0

    def test_multiple_crew_and_passengers(self, registry, resolver):
        mapping = registry.get_form("LSGS", "lsgs")
        request = GenerateRequest(
            airport="LSGS",
            form="lsgs",
            flight=make_flight(origin="ZZZZ", destination="LSGS"),
            aircraft=make_aircraft(),
            crew=[make_pilot(), make_crew_member()],
            passengers=[make_passenger(), make_second_passenger()],
        )
        result = fill_pdf(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0


# ── French customs PDF ────────────────────────────────────────────────────────

class TestFrenchCustomsFiller:
    @pytest.fixture
    def registry(self):
        return MappingRegistry(str(MAPPINGS_DIR), str(TEMPLATES_DIR))

    @pytest.fixture
    def resolver(self):
        return StubAirportResolver()

    def _generate(self, registry, resolver, flatten=False) -> bytes:
        mapping = registry.get_form("LFAC", "french_customs")
        request = GenerateRequest(
            airport="LFAC",
            form="french_customs",
            flight=make_flight(origin="ZZZZ", destination="LFAC"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
            passengers=[make_passenger(), make_second_passenger()],
        )
        return fill_french_customs(
            registry.get_template_path(mapping),
            mapping,
            request,
            resolver,
            flatten=flatten,
        )

    def test_output_is_valid_pdf(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        assert len(reader.pages) >= 1

    def test_flattened(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, flatten=True)
        reader = PdfReader(BytesIO(pdf_bytes))
        # Flattened PDF should have no widget annotations on pages
        for page in reader.pages:
            annots = page.get("/Annots", [])
            assert len(annots) == 0

    def test_combined_person_list(self, registry, resolver):
        """Crew + passengers should be combined — 1 crew + 2 pax = 3 people."""
        mapping = registry.get_form("LFAC", "french_customs")
        request = GenerateRequest(
            airport="LFAC",
            form="french_customs",
            flight=make_flight(origin="ZZZZ", destination="LFAC"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
            passengers=[make_passenger(), make_second_passenger()],
        )
        # Just verify it generates without error for 3 people
        result = fill_french_customs(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0

    def test_departure_direction(self, registry, resolver):
        mapping = registry.get_form("LFAC", "french_customs")
        request = GenerateRequest(
            airport="LFAC",
            form="french_customs",
            flight=make_flight(origin="LFAC", destination="ZZZZ"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
        )
        result = fill_french_customs(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0


# ── PDF filler (LFQA customs) ────────────────────────────────────────────────

class TestPdfFillerLFQA:
    @pytest.fixture
    def registry(self):
        return MappingRegistry(str(MAPPINGS_DIR), str(TEMPLATES_DIR))

    @pytest.fixture
    def resolver(self):
        return StubAirportResolver()

    def _generate(self, registry, resolver, origin="ZZZZ", destination="LFQA") -> bytes:
        mapping = registry.get_form("LFQA", "lfqa")
        request = GenerateRequest(
            airport="LFQA",
            form="lfqa",
            flight=make_flight(origin=origin, destination=destination),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
            passengers=[make_passenger()],
            observations="Fictional test observation",
        )
        return fill_pdf(
            registry.get_template_path(mapping), mapping, request, resolver
        )

    def test_output_is_valid_pdf(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        assert len(reader.pages) >= 1

    def test_arrival_mark(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, origin="ZZZZ", destination="LFQA")
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert fields["DEMANDE DARRIVEE"].get("/V") == "X"
        assert fields["DEMANDE DE DEPART"].get("/V") in ("", None)

    def test_departure_mark(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, origin="LFQA", destination="ZZZZ")
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert fields["DEMANDE DE DEPART"].get("/V") == "X"
        assert fields["DEMANDE DARRIVEE"].get("/V") in ("", None)

    def test_arrival_fills_arrival_side_only(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, origin="ZZZZ", destination="LFQA")
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert "ZZ-TST" in str(fields["ARRIVALAIRCRAFTREG"].get("/V", ""))
        assert fields.get("DEPARTUREAIRCRAFTREG", {}).get("/V") is None

    def test_departure_fills_departure_side_only(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver, origin="LFQA", destination="ZZZZ")
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert "ZZ-TST" in str(fields["DEPARTUREAIRCRAFTREG"].get("/V", ""))
        assert fields.get("ARRIVALAIRCRAFTREG", {}).get("/V") is None

    def test_crew_filled(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert str(fields["NOM NAMERow1"].get("/V", "")) == "Kowalski"
        assert str(fields["PRENOMRow1"].get("/V", "")) == "Zara"

    def test_passenger_filled(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert str(fields["NOM NAMERow1_2"].get("/V", "")) == "Petrova"
        assert str(fields["PRENOMRow1_2"].get("/V", "")) == "Lina"

    def test_observations_filled(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert str(fields["OBSERVATIONS"].get("/V", "")) == "Fictional test observation"

    def test_airport_icao_filled(self, registry, resolver):
        pdf_bytes = self._generate(registry, resolver)
        reader = PdfReader(BytesIO(pdf_bytes))
        fields = reader.get_fields() or {}
        assert str(fields["AIRPORTICAO"].get("/V", "")) == "LFQA"


# ── XLSX filler (GAR) ─────────────────────────────────────────────────────────

class TestXlsxFiller:
    @pytest.fixture
    def registry(self):
        return MappingRegistry(str(MAPPINGS_DIR), str(TEMPLATES_DIR))

    @pytest.fixture
    def resolver(self):
        return StubAirportResolver()

    def _generate(self, registry, resolver) -> bytes:
        mapping = registry.get_form("EGKA", "gar")
        request = GenerateRequest(
            airport="EGKA",
            form="gar",
            flight=make_flight(origin="ZZZZ", destination="EGKA"),
            aircraft=make_aircraft(),
            crew=[make_pilot(), make_crew_member()],
            passengers=[make_passenger()],
            extra_fields={
                "reason_for_visit": "Based",
                "responsible_person": {"name": "Zara Kowalski", "address": "7 Birch Lane"},
            },
        )
        return fill_xlsx(
            registry.get_template_path(mapping), mapping, request, resolver
        )

    def test_output_is_valid_xlsx(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "GAR" in wb.sheetnames

    def test_header_direction(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        assert ws["B2"].value == "ARRIVAL"

    def test_header_departure_direction(self, registry, resolver):
        mapping = registry.get_form("EGKA", "gar")
        request = GenerateRequest(
            airport="EGKA",
            form="gar",
            flight=make_flight(origin="EGKA", destination="ZZZZ"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
            extra_fields={
                "reason_for_visit": "Based",
                "responsible_person": {"name": "Zara Kowalski", "address": "7 Birch Lane"},
            },
        )
        xlsx_bytes = fill_xlsx(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        assert ws["B2"].value == "DEPARTURE"

    def test_registration_filled(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        assert ws["B5"].value == "ZZ-TST"

    def test_crew_rows(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        # Crew starts at row 9, columns E=last_name, F=first_name
        assert ws["E9"].value == "Kowalski"
        assert ws["F9"].value == "Zara"
        assert ws["E10"].value == "Bergström"
        assert ws["F10"].value == "Tariq"

    def test_passenger_rows(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        # Passengers start at row 20
        assert ws["E20"].value == "Petrova"
        assert ws["F20"].value == "Lina"

    def test_extra_field_choice(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        assert ws["B6"].value == "Based"

    def test_extra_field_person_address(self, registry, resolver):
        xlsx_bytes = self._generate(registry, resolver)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb["GAR"]
        # responsible_person address maps to D6
        assert ws["D6"].value == "7 Birch Lane"

    def test_no_extra_fields(self, registry, resolver):
        """Generation should work even without extra fields."""
        mapping = registry.get_form("EGKA", "gar")
        request = GenerateRequest(
            airport="EGKA",
            form="gar",
            flight=make_flight(origin="ZZZZ", destination="EGKA"),
            aircraft=make_aircraft(),
            crew=[make_pilot()],
        )
        result = fill_xlsx(
            registry.get_template_path(mapping), mapping, request, resolver
        )
        assert len(result) > 0

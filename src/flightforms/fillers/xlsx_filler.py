"""XLSX filler for spreadsheet-based forms (GAR, myhandling, etc.)."""

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from ..api.models import GenerateRequest
from ..registry import FormMapping


def _parse_date(date_str: str) -> datetime:
    """Parse ISO date string and return a datetime for Excel native date cells."""
    return datetime.strptime(date_str, "%Y-%m-%d")


def _format_date(date_str: str, fmt: str) -> str:
    """Parse ISO date string and return formatted string (e.g. '03/04/2026')."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return dt.strftime(fmt)


def _format_time(time_str: str, fmt: str) -> str:
    """Format HH:MM to HH:MM:SS if needed."""
    if fmt == "HH:MM:SS" and len(time_str) == 5:
        return time_str + ":00"
    return time_str


def _load_fbo_lookup(mapping: FormMapping) -> dict[str, str]:
    """Load ICAO → FBO ID lookup from the fbo_lookup file referenced in mapping."""
    fbo_file = mapping.raw.get("fbo_lookup")
    if not fbo_file:
        return {}
    mappings_dir = Path(__file__).parent.parent / "mappings"
    path = (mappings_dir / fbo_file).resolve()
    if not path.is_relative_to(mappings_dir.resolve()):
        raise ValueError("Invalid fbo_lookup path")
    if not path.exists():
        return {}
    with open(path) as f:
        return json.load(f)


def fill_xlsx(
    template_path: Path,
    mapping: FormMapping,
    request: GenerateRequest,
    airport_resolver,
) -> bytes:
    wb = load_workbook(str(template_path))
    sheet_name = mapping.raw.get("sheet", "GAR")
    ws = wb[sheet_name]

    header_map = mapping.raw.get("header_map", {})
    person_columns = mapping.raw.get("person_columns", {})
    crew_start = mapping.raw.get("crew_start_row", 9)
    pax_start = mapping.raw.get("pax_start_row", 20)
    date_fmt = mapping.raw.get("date_format", "DD/MM/YYYY")
    time_fmt = mapping.raw.get("time_format", "HH:MM:SS")

    is_arrival = request.airport == request.flight.destination

    # Header fields
    direction_text = "ARRIVAL" if is_arrival else "DEPARTURE"
    captain = request.crew[0].last_name if request.crew else ""
    observations = request.observations or mapping.default_observations or ""

    header_values = {
        "direction": direction_text,
        "arrival_icao": request.flight.destination,
        "arrival_date": _parse_date(request.flight.arrival_date),
        "arrival_time": _format_time(request.flight.arrival_time_utc, time_fmt),
        "departure_icao": request.flight.origin,
        "departure_date": _parse_date(request.flight.departure_date),
        "departure_time": _format_time(request.flight.departure_time_utc, time_fmt),
        "owner": request.aircraft.owner or "",
        "contact": request.flight.contact or "",
        "registration": request.aircraft.registration,
        "aircraft_type": request.aircraft.type,
        "captain_surname": captain,
        "usual_base": request.aircraft.usual_base or "",
    }

    for field_name, cell_ref in header_map.items():
        if field_name in header_values:
            ws[cell_ref] = header_values[field_name]

    # Extra fields (text/choice → header_map cell, person → maps_to cells)
    for ef in mapping.extra_fields:
        ef_def = ef if isinstance(ef, dict) else {"key": ef, "type": "text"}
        key = ef_def["key"]
        ef_type = ef_def.get("type", "text")
        value = (request.extra_fields or {}).get(key)

        if ef_type in ("text", "choice"):
            if key in header_map and value:
                ws[header_map[key]] = value
        elif ef_type == "person" and isinstance(value, dict):
            # Person name → header_map cell if present
            if key in header_map:
                ws[header_map[key]] = value.get("name", "")
            # Sub-fields → maps_to cells
            for sub_field, cell_ref in ef_def.get("maps_to", {}).items():
                ws[cell_ref] = value.get(sub_field, "")

    # Fill crew
    for i, crew in enumerate(request.crew):
        row = crew_start + i
        _fill_person_row(ws, row, person_columns, crew, date_fmt)

    # Fill passengers
    for i, pax in enumerate(request.passengers):
        row = pax_start + i
        _fill_person_row(ws, row, person_columns, pax, date_fmt)

    # Column map: fill a single data row with values keyed by column letter.
    # Used by forms like myhandling where one row = one flight movement.
    column_map = mapping.raw.get("column_map")
    if column_map:
        data_row = mapping.raw.get("data_start_row", 4)
        fbo_lookup = _load_fbo_lookup(mapping)

        # Flight type mapping (nature string → form enum value)
        flight_type_map = mapping.raw.get("flight_type_map", {})
        flight_type = flight_type_map.get(
            request.flight.nature.lower(),
            request.flight.nature,
        )

        # Build values for the arrival (main flight) side
        values = {
            "arrival_date": _format_date(request.flight.arrival_date, date_fmt),
            "arrival_time": _format_time(request.flight.arrival_time_utc, time_fmt),
            "registration": request.aircraft.registration,
            "aircraft_type": request.aircraft.type,
            "mtow": "",
            "arrival_flight_number": "",
            "arrival_flight_type": flight_type,
            "arrival_from": request.flight.origin,
            "pax_in": str(len(request.passengers)),
            "crew_in": str(len(request.crew)),
            "bags_in": "",
            "fbo_location": request.airport,
        }

        # FBO ID: look up from fbo_lookup, format as "id|name" or just ICAO
        fbo_id = fbo_lookup.get(request.airport)
        if fbo_id:
            values["fbo_id"] = fbo_id

        # Departure side from connecting flight
        if request.connecting_flight:
            cf = request.connecting_flight
            dep_flight_type = flight_type_map.get(
                request.flight.nature.lower(),
                request.flight.nature,
            )
            values.update({
                "departure_date": _format_date(cf.departure_date, date_fmt),
                "departure_time": _format_time(cf.departure_time_utc, time_fmt),
                "departure_flight_number": "",
                "departure_flight_type": dep_flight_type,
                "departure_to": cf.destination,
                "pax_out": str(len(request.passengers)),
                "crew_out": str(len(request.crew)),
                "bags_out": "",
            })

        for field_name, col_letter in column_map.items():
            if field_name in values and values[field_name]:
                ws[f"{col_letter}{data_row}"] = values[field_name]

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _fill_person_row(ws, row: int, columns: dict, person, date_fmt: str):
    """Fill a single person row in the spreadsheet."""
    field_values = {
        "id_type": person.id_type or "",
        "id_type_other": "",
        "id_issuing_country": person.id_issuing_country or "",
        "id_number": person.id_number or "",
        "last_name": person.last_name,
        "first_name": person.first_name,
        "sex": person.sex or "",
        "dob": _parse_date(person.dob) if person.dob else "",
        "place_of_birth": person.place_of_birth or "",
        "nationality": person.nationality or "",
        "id_expiry": _parse_date(person.id_expiry) if person.id_expiry else "",
        "address": "",
    }

    for field_name, col_letter in columns.items():
        if field_name in field_values:
            ws[f"{col_letter}{row}"] = field_values[field_name]
